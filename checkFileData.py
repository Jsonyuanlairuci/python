from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from collections import OrderedDict

requestDomain="https://filscan.io/address/miner?address="
nodesArr={'节点ID','节点ID'}


excelRow=2           #todo 每日需要修改该数字，以便修改插入的列数

# 请求网站获取所需数据，并把数据打包返回
def getNodeDtaReponse(nodeId):
    requestUrl=requestDomain+nodeId
    print(requestUrl)
    driver=webdriver.Chrome()
    driver.get(requestUrl)
    data={}
    try:
        data['fil_balance'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "num"))
        ).text
        data['available'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "available"))
        ).text
        data['pledged'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "pledged"))
        ).text
        data['deposits'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "deposits"))
        ).text
        data['reward'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "reward"))
        ).text
        data['power'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[@class="left font-20 font-400 text flex align-center"]'))
        ).text
        data['total_reward'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(),"总奖励: ")]/following-sibling::div[1]'))
        ).text
        data['total_all_status'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(),"扇区状态: ")]/following-sibling::div[1]/span'))
        ).text
        data['total_used_status'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(),"扇区状态: ")]/following-sibling::div[1]/span[@class="proving"]'))
        ).text
        data['total_fault_status'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(),"扇区状态: ")]/following-sibling::div[1]/span[@class="fault"]'))
        ).text
        data['total_pre_status'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(),"扇区状态: ")]/following-sibling::div[1]/span[@class="pre"]'))
        ).text
        data['power_add_num'] = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//div[@class="content-growth"]/div[@class="growth-item"]/div[@class="value"]'))
        ).text
        return data
    finally:
        driver.quit()
    pass

# 循环所有节点ID,    由于python无序循环，所以只能单独处理循环的顺序问题
def cycleNode():
    nodeData={}
    od= OrderedDict()
    od[0] = '节点ID'

    for key,node in od.items():
        nodeInfo=getNodeDtaReponse(node)
        nodeData[node]=nodeInfo
        nodeData[node]['current_time']=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    writeExcel(nodeData)
    pass

from openpyxl import load_workbook


# 将数据写入excel表格
def writeExcel(data):
    wb=load_workbook(filename='/home/wmkj/桌面/项目/fil/节点数据/2022-4.xlsx')  #todo  每个月需要修改一次
    sheet=wb.sheetnames
    writeSheet=wb[sheet[0]]
    nodeIndex=0
    dateIndex=2
    print(data)
    for key,row in data.items():
        excelColumn=len(row)*nodeIndex+dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=key)                    #节点ID
        print([excelRow,excelColumn])
        dateIndex+=1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['current_time'])             #查询时间
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['fil_balance'])              #余额
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['available'])                #可用余额
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['pledged'])                   #扇区质押
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['deposits'])                 #预存款
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['reward'])                   #锁仓奖励
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['power'])                    #有效算力
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['total_reward'])            #总奖励
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['total_all_status'])         #扇区状态
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['total_used_status'])         #有效扇区
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['total_fault_status'])         #错误扇区
        print([excelRow, excelColumn])
        dateIndex += 1
        excelColumn = len(row) * nodeIndex + dateIndex
        writeSheet.cell(row=excelColumn, column=excelRow, value=row['power_add_num'])             #算力增量
        print([excelRow, excelColumn])
        nodeIndex+=1
        dateIndex = 2+nodeIndex
    wb.save(r"/home/wmkj/桌面/项目/fil/节点数据/2022-4.xlsx")
    pass


def main():
    cycleNode()

    pass

if __name__=="__main__":
    main()